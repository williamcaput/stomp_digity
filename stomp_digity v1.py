#!python
####################################################################################################################
#  Stomp_digity.py   - Gimme two pair....keep stomp'en in those air force ones
#
#  Version:          0.1.1
#
#  Author:           Me, Sam Reinthaler, Fran 'Danger' Brown (Perl Version) 
#
#  Last Revision:    10/10/2015
#
#  Description:      Stomp_digity.py takes in the name of an Nmap grepable result file and (optionally) the name of
#                    the desired excel .xls output file. It parses the portscan file, extracts information about
#                    open ports, OS guesses, FQDN, etc. and then populates the excel spreadsheet.
#                    Perfect for easy cut-n-pasting into reports.
#
#  Changes > v0.1.1:   //
#
#  Tested on:        Nmap 6.40 grepable portscan files
#
#  Tested using:     Python 3.4.3 on Linux, OpenPyXL 2.2.6, LibreOffice Calc 4.2.8.2
#
#  Usage:            python Stomp_digity.py <nmap_portscan_result_file> [-o <outputfile.xlsx>]
# 
#####################################################################################################################

#  Interprets Nmap grepable output                  https://nmap.org/book/output-formats-grepable-output.html
#  Outputs to .xlsx with OpenPyXL (Python 3 only)   http://openpyxl.readthedocs.org/en/latest/index.html



###################
# ENABLE MODULES  #
###################

import argparse
import re # regex
import os # path.exists and path.splitext
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color



##############
#  SETTINGS  #
##############

#  you can add/remove/change ports as long as the HEADERS and COLUMN_WIDTHS arrays match. COMMON PORTS derives port numbers from HEADERS unless explicitly set.
HEADERS = [ "IP Address", "Fully-qualified\nDomain Name", "Operating System Guess", "FTP (21)", "SSH (22)", "Telnet (23)", "SMTP (25)", "Domain (53)", "HTTP (80)", "HTTPS (443)", "Other" ]
COMMON_PORTS = [ re.match(r'[^\(]+\((\d+)\)', p).group(1) for p in HEADERS[3:-1] ]   #  derive port numbers from the HEADERS array
#COMMON_PORTS = [21,22,23,25,53,80,443]                                              #  uncomment this line and comment line above to set port numbers manually
COLUMN_WIDTHS = [20.5, 25.9, 26.38] + [3]*7 + [21.5]                                 #  [20.5, 28, 26.38, 3, 3, 3, 3, 3, 3, 3, 21.5]
HEADER_ROW_HEIGHT = 80
COLOR_HEADER_TEXT   = 'FFFFFF'
COLOR_HEADER_FILL   = '004C98'
COLOR_HEADER_BORDER = '00800000'

num_common_ports = len(COMMON_PORTS)



#############
#  HELPERS  #
#############

#  OUTPUT A PROMPT AND WAIT FOR Y/N
def query_yes_no(question, default="yes"):

   valid = {"yes": True, "y": True, "no": False, "n": False}
   if default not in ['yes','no',None]:
      raise ValueError("invalid default answer: '%s'" % default)
   prompt = ' [' + ('Y' if default is 'yes' else 'y') + '/' + ('N' if default is 'no' else 'n') + '] '

   while True:
      choice = input(question + prompt).lower()
      if default is not None and choice == '':
         return valid[default]
      elif choice in valid:
         return valid[choice]
      else:
         print("Please respond with 'yes', 'no', 'y', or 'n').")

#  CREATE IP TUPLE KEY FOR SORTING HOSTS BY IP
#  http://stackoverflow.com/questions/6545023/how-to-sort-ip-addresses-stored-in-dictionary-in-python/6545090#6545090
def ip_key(host):
    return tuple(int(part) for part in host['ip'].split('.'))



##################
#  PROCESS ARGS  #
##################

def process_args():

   # PARSE COMMAND-LINE ARGUMENTS
   parser = argparse.ArgumentParser()
   parser.add_argument('--output_file', '-o')
   parser.add_argument("input_file")
   args = parser.parse_args()
   input_file = args.input_file
   output_file = args.output_file

   #  check that input file exists
   if not os.path.isfile(input_file):
      print('Input file does not exist.')
      return (True, None, None)
 
   #  if no output file given, create from input file name and tell user
   if output_file is None:
      output_file = os.path.splitext(input_file)[0] + '.xlsx'
      print('Output filename generated from input filename: ',output_file)

   #  check correct file extension(s) (warning?)
   if (not output_file.endswith('.xlsx')) and (not query_yes_no('Output file not .xlsx. Continue?',default='no')):
      return (True, None, None) #  exit

   #  if output file exists, warn?
   if (os.path.exists(output_file)) and (not query_yes_no('Output file exists. Overwrite?',default='no')):
      return (True, None, None) #  exit

   return (False, input_file, output_file)



###################
#  PROCESS GNMAP  #
###################

def process_gnmap(filename):

   hosts = []

   with open(filename,'r') as input_file:
      for line in input_file:

         if (not line.startswith('Host: ')) or ('/open/' not in line):     #  if line is not a host or has no open ports, skip it
            continue
   
         host = {}

         fields = line.split('\t')                                         # fields (sections) are separated by tab character, and start with field name, then ": " 

         for field in fields:

            if field.startswith('Host: '):
               host_field = field.split()
               host['ip'] = host_field[1]
               host['fqdn'] = host_field[2].strip('()')

            elif field.startswith('Ports: '):
               matches = re.finditer(r' (\d+)\/open\/', field)             #  creates iterator of match objects for all matches in string
               open_ports = [match.group(1) for match in matches]          #  group(0) = whole match, group(1) = 1st group inside (just port #)
               host['common_ports'] = [('X' if p in open_ports else '') for p in COMMON_PORTS] #  for every common port, put 'X' if open in current host, and '' if not
               host['other_ports'] = [p for p in open_ports if p not in COMMON_PORTS]          #  all open ports for current host that are not common ports

            elif field.startswith('OS: '):
               host['os'] = field[4:]                                      #  remove 'OS: ' and take rest of field

         hosts.append(host)

   hosts.sort(key=ip_key)                                                  #  Sort hosts by ip address

   return hosts



############################
#  PRINT HEADERS AND DATA  #
############################

def print_headers(ws):

   #  SET HEADER CELL DATA
   for col in range(len(HEADERS)):
      ws.cell(row=1, column=col+1).value = HEADERS[col]

def print_hosts(ws, hosts):

   for h in range(len(hosts)):

      #  IP, FQDN, OS
      ws.cell(row=h+2, column=1).value = hosts[h]['ip']
      ws.cell(row=h+2, column=2).value = hosts[h]['fqdn']
      ws.cell(row=h+2, column=3).value = hosts[h].get('os', '')            #  get() takes a default; some hosts don't have os key

      #  COMMON PORTS
      for p in range(num_common_ports):
         ws.cell(row=h+2, column=4+p).value = hosts[h]['common_ports'][p]

      #  OTHER PORTS
      ws.cell(row=h+2, column=4+num_common_ports).value = ', '.join(hosts[h]['other_ports'])



################
#  FORMATTING  #
################

def format_cell_dimensions(ws):                                            #  best_fit doesn't work; setting column dimensions manually
   ws.row_dimensions[1].height = HEADER_ROW_HEIGHT                         #  default = 80
   for col in range(len(COLUMN_WIDTHS)):                                   #  default = [20.5, 25.9, 26.38, 3, 3, 3, 3, 3, 3, 3, 21.5]
      ws.column_dimensions[chr(col+65)].width = COLUMN_WIDTHS[col]         #  chr(col+65) gives column name (A = 65 ASCII)

def format_headers(ws):






   #  CREATE HEADER CELL FORMATS
   fmt_h_font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TEXT)
   fmt_h_fill = PatternFill(fill_type='solid', start_color=COLOR_HEADER_FILL)
   fmt_h_border = Border(outline=Side(border_style=None, color=COLOR_HEADER_BORDER))
   fmt_hmain_align = Alignment(horizontal='center', text_rotation=0, wrap_text=True, indent=0)
   fmt_hport_align = Alignment(horizontal='center', text_rotation=90, wrap_text=True, indent=0)

   #  APPLY HEADER CELL FORMATS
   for col in range(len(HEADERS)):
      c = ws.cell(row=1, column=col+1)
      (c.font, c.fill, c.border) = (fmt_h_font, fmt_h_fill, fmt_h_border)
      c.alignment = fmt_hmain_align if (col not in range(3, len(HEADERS)-1)) else fmt_hport_align

def format_hosts(ws, num_hosts):

   fmt_os_align = Alignment(wrap_text=True)
   fmt_other_port_align = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
   fmt_common_port_align = Alignment(horizontal='center')

   for r in range(num_hosts):

      #  WrapText the "OS Guess" and "Other Ports" columns, and align 'Other Ports' column to bottom-left
      ws.cell(row=2+r, column=3).alignment = fmt_os_align
      ws.cell(row=2+r, column=num_common_ports+4).alignment = fmt_other_port_align

      #  Center the X's for "Common Ports" columns (i.e. ftp, telnet,..,https)
      for c in range(num_common_ports):
         ws.cell(row=2+r, column=c+4).alignment = fmt_common_port_align



##########
#  MAIN  #
##########

def main():

   # ARGUMENT PARSING, ERROR CHECKING, AND CONFIRMATION
   (input_error, input_file, output_file) = process_args()
   if input_error is True:
      print('Exiting.')
      return 1
   
   hosts = process_gnmap(input_file)   #  get data from nmap output file

   wb = Workbook()                     #  create a new workbook
   ws = wb.active                      #  get active worksheet
   ws.title = "Target Matrix"          #  rename sheet "Target Matrix"

   print_headers(ws)                   #  print headers
   print_hosts(ws, hosts)              #  print data

   format_cell_dimensions(ws)          #  format cells
   format_headers(ws)                  #
   format_hosts(ws, len(hosts))        #


   #  WRITE FILE
   wb.save(filename = output_file)
   print("File written.")
   return 0



if __name__ == '__main__':
    main()